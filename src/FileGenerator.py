from openpyxl import load_workbook
import os.path

from ExcelCodeName import ExcelCodeName

__author__ = "Au Wai Lun Andy"
__version__ = "0.0.1"
__status__ = "Test"


class FileGenerator:

    def __init__(self, filename: str, sheetname: str):
        self.workbook = load_workbook(filename, read_only=False)
        self.worksheet = self.workbook[sheetname]

    def getWorkbook(self):
        return self.workbook

    def getWorksheet(self):
        return self.worksheet

    def generateCsv(self, outputFilename: str, headerStart: ExcelCodeName, headerFieldCount: int, detailStart: ExcelCodeName, detailCount: int, detailFieldCount: int, trailerStart: ExcelCodeName, trailerFieldCount: int):

        # CSV file
        csv_file = open(outputFilename, "w")

        # generate header line
        header: str = ""
        if headerFieldCount > 0:
            temp = headerStart
            for i in range(0, headerFieldCount):
                if i == 0:
                    header = header + "\""
                else:
                    header = header + ",\""
                header = header + str(self.worksheet[str(temp)].value) + "\""
                temp.nextColumn()

        if len(header) > 0:
            csv_file.write(header + "\n")

        # generate detail lines
        details = []
        if detailCount > 0 and detailFieldCount > 0:
            temp = detailStart
            for l in range(0, detailCount):
                print("temp = {}".format(temp.getCode()))
                detail = ""
                temp2 = ExcelCodeName(temp.getCode())
                for i in range(0, detailFieldCount):
                    print("temp2 = {}".format(temp2.getCode()))
                    if i == 0:
                        detail = detail + "\""
                    else:
                        detail = detail + ",\""
                    detail = detail + \
                        self.worksheet[temp2.getCode()].value + "\""
                    temp2.nextColumn()
                details.append(detail)
                temp.nextRow()

        if len(details) > 0:
            for detail in details:
                csv_file.write(detail + "\n")

        # generate trailer line
        trailer: str = ""
        if trailerFieldCount > 0:
            temp = trailerStart
            for i in range(0, trailerFieldCount):
                if i == 0:
                    trailer = trailer + "\""
                else:
                    trailer = trailer + ",\""
                trailer = trailer + str(self.worksheet[str(temp)].value) + "\""
                temp.nextColumn()

        if len(trailer) > 0:
            csv_file.write(trailer)

        csv_file.close()

    def generateDat(self, outputFilename: str, headerStart: ExcelCodeName, headerFieldCount: int, headerLengthOffset: int, detailStart: ExcelCodeName, detailCount: int, detailFieldCount: int, detailLengthOffset: int, trailerStart: ExcelCodeName, trailerFieldCount: int, trailerLengthOffset: int):
        pass


if __name__ == "__main__":

    print("begin")

    # TODO
    # move to config later
    dir = "/../"
    filename = "Book1.xlsx"
    path = os.path.dirname(__file__) + dir + filename
    sheetname = "DDI"

    # def generateCsv(self, outputFilename: str, headerStart: ExcelCodeName, headerFieldCount: int, detailStart: ExcelCodeName, detailFieldCount: int, trailerStart: ExcelCodeName, trailerFieldCount: int):

    generator = FileGenerator(path, sheetname)
    generator.generateCsv(sheetname + ".CSV", ExcelCodeName("B8"),
                          5, ExcelCodeName("B17"), 2, 8, ExcelCodeName("B22"), 3)

    print("end")
    pass

    # wb = load_workbook(filename, read_only=False)
    # ws = wb.create_sheet(sheetname+"_DATA")
    # ws["A16"] = script
    # for sheet in wb.worksheets:
    #     print(sheet.title)

    # wb.save("Book.xlsx")
